import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import random
from pathlib import Path
from typing import Optional


class ExcelBebanProcessor:
    """
    Processor untuk menghitung nilai BEBAN pada file Excel
    berdasarkan MUTU, KODE_BENDA_UJI, UMUR, dan DOCKET
    """
    
    def __init__(self, file_path: str, sheet_name: str = 'ODOO'):
        """
        Inisialisasi processor
        
        Args:
            file_path: Path ke file Excel
            sheet_name: Nama sheet yang akan diproses (default: 'ODOO')
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None
        
        # Tracking untuk logika umur 28
        self.previous_docket: Optional[str] = None
        self.previous_beban: Optional[float] = None
        
        # Mapping kolom (1-based index untuk Excel)
        self.column_map = {
            'DOCKET': 1,      # Kolom A -> index 1
            'MUTU': 6,        # Kolom F -> index 6
            'UMUR': 9,        # Kolom I -> index 9
            'KODE_BENDA_UJI': 16,  # Kolom P -> index 16
            'BEBAN': 18       # Kolom R -> index 18
        }
        
        # Keywords untuk pengecekan
        self.keywords = ["PP - TOL PTB", "WASKITA - ABP JO", "HK - JAKON JO", "WIKA - ADHI JO"]
    
    def load_excel(self) -> bool:
        """Membuka file Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook[self.sheet_name]
            print(f"✓ File Excel berhasil dibuka: {self.file_path}")
            print(f"✓ Sheet aktif: {self.sheet_name}")
            return True
        except FileNotFoundError:
            print(f"✗ Error: File tidak ditemukan - {self.file_path}")
            return False
        except KeyError:
            print(f"✗ Error: Sheet '{self.sheet_name}' tidak ditemukan")
            return False
        except Exception as e:
            print(f"✗ Error saat membuka file: {str(e)}")
            return False
    
    def get_cell_value(self, row_index: int, column_key: str):
        """Mendapatkan nilai cell berdasarkan row dan column key"""
        if self.sheet is None:
            raise RuntimeError("Sheet belum dimuat. Jalankan load_excel() terlebih dahulu.")
        
        col_index = self.column_map[column_key]
        return self.sheet.cell(row=row_index, column=col_index).value
    
    def set_cell_value(self, row_index: int, column_key: str, value: str) -> None:
        """Mengisi nilai cell berdasarkan row dan column key"""
        if self.sheet is None:
            raise RuntimeError("Sheet belum dimuat. Jalankan load_excel() terlebih dahulu.")
        
        col_index = self.column_map[column_key]
        self.sheet.cell(row=row_index, column=col_index, value=value)
    
    def calculate_beban(self, row_index: int) -> Optional[str]:
        """
        Menghitung nilai BEBAN untuk baris tertentu
        
        Args:
            row_index: Index baris di Excel (1-based)
            
        Returns:
            String nilai BEBAN atau None jika tidak memenuhi kriteria
        """
        # Ambil data dari row
        mutu = str(self.get_cell_value(row_index, 'MUTU') or '')
        kode_benda_uji = str(self.get_cell_value(row_index, 'KODE_BENDA_UJI') or '')
        umur = self.get_cell_value(row_index, 'UMUR')
        docket = self.get_cell_value(row_index, 'DOCKET')


        if "Class B-2" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(390.12, 460.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(580, 700)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (580 + 700) / 2  # 640
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(580, min(self.previous_beban + adjustment, 700))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (580 + 700) / 2  # 640
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-59, -40)  # Turun signifikan
                    else:
                        adjustment = random.uniform(40, 59)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(580, min(new_beban, 700))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(580, 700)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "K-400" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(390.12, 460.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(580, 700)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (580 + 700) / 2  # 640
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(580, min(self.previous_beban + adjustment, 700))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (580 + 700) / 2  # 640
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-59, -40)  # Turun signifikan
                    else:
                        adjustment = random.uniform(40, 59)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(580, min(new_beban, 700))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(580, 700)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Class B-1" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(360.12, 420.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(540, 625)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (540 + 625) / 2  # 582.5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(540, min(self.previous_beban + adjustment, 625))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (540 + 625) / 2  # 582.5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-42, -30)  # Turun signifikan
                    else:
                        adjustment = random.uniform(30, 42)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(540, min(new_beban, 625))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(540, 625)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "Fc-30" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(360.12, 420.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(540, 625)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (540 + 625) / 2  # 582.5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(540, min(self.previous_beban + adjustment, 625))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (540 + 625) / 2  # 582.5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-42, -30)  # Turun signifikan
                    else:
                        adjustment = random.uniform(30, 42)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(540, min(new_beban, 625))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(540, 625)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "K-300" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(290.12, 360.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(430, 500)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (430 + 500) / 2  # 465
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(430, min(self.previous_beban + adjustment, 500))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (430 + 500) / 2  # 465
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-34, -20)  # Turun signifikan
                    else:
                        adjustment = random.uniform(20, 34)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(430, min(new_beban, 500))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(430, 500)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Fc-25" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(290.12, 360.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(430, 500)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (430 + 500) / 2  # 465
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(430, min(self.previous_beban + adjustment, 500))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (430 + 500) / 2  # 465
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-34, -20)  # Turun signifikan
                    else:
                        adjustment = random.uniform(20, 34)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(430, min(new_beban, 500))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(430, 500)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "K-250" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(230.12, 285.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(355, 410)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (355 + 410) / 2  # 382,5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(355, min(self.previous_beban + adjustment, 410))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (355 + 410) / 2  # 382,5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-26, -11)  # Turun signifikan
                    else:
                        adjustment = random.uniform(11, 26)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(355, min(new_beban, 410))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(355, 410)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Fc-20" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(230.12, 285.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(355, 410)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (355 + 410) / 2  # 382,5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(355, min(self.previous_beban + adjustment, 410))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (355 + 410) / 2  # 382,5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-26, -11)  # Turun signifikan
                    else:
                        adjustment = random.uniform(11, 26)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(355, min(new_beban, 410))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(355, 410)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Class C" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(230.12, 285.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(355, 410)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (355 + 410) / 2  # 382,5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(355, min(self.previous_beban + adjustment, 410))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (355 + 410) / 2  # 382,5
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-26, -11)  # Turun signifikan
                    else:
                        adjustment = random.uniform(11, 26)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(355, min(new_beban, 410))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(355, 410)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "K-175" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(165.12, 215.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(255, 305)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(255, min(self.previous_beban + adjustment, 305))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-24, -8)  # Turun signifikan
                    else:
                        adjustment = random.uniform(8, 24)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(255, min(new_beban, 305))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(255, 305)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Fc-15" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(165.12, 215.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(255, 305)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(255, min(self.previous_beban + adjustment, 305))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-24, -8)  # Turun signifikan
                    else:
                        adjustment = random.uniform(8, 24)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(255, min(new_beban, 305))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(255, 305)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Class D" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(165.12, 215.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(255, 305)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(255, min(self.previous_beban + adjustment, 305))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-24, -8)  # Turun signifikan
                    else:
                        adjustment = random.uniform(8, 24)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(255, min(new_beban, 305))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(255, 305)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"

        elif "K-125" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(120.12, 165.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(180, 240)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (180 + 240) / 2  # 210
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(180, min(self.previous_beban + adjustment, 240))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (180 + 240) / 2  # 210
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -12)  # Turun signifikan
                    else:
                        adjustment = random.uniform(12, 29)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(180, min(new_beban, 240))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(180, 240)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "Fc-10" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(120.12, 165.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(180, 240)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (180 + 240) / 2  # 210
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(180, min(self.previous_beban + adjustment, 240))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (180 + 240) / 2  # 210
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -12)  # Turun signifikan
                    else:
                        adjustment = random.uniform(12, 29)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(180, min(new_beban, 240))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(180, 240)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"

        elif "Class E" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(120.12, 165.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(180, 240)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (180 + 240) / 2  # 210
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(180, min(self.previous_beban + adjustment, 240))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (180 + 240) / 2  # 210
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -12)  # Turun signifikan
                    else:
                        adjustment = random.uniform(12, 29)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(180, min(new_beban, 240))
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(180, 240)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"

        else:
            return None



    def process_all_rows(self, start_row: int = 2) -> None:
        """
        Memproses semua baris dalam sheet
        
        Args:
            start_row: Baris awal untuk diproses (default: 2, skip header)
        """
        if self.sheet is None:
            print("✗ Sheet belum dimuat. Jalankan load_excel() terlebih dahulu.")
            return
        
        total_rows = self.sheet.max_row
        if total_rows is None:
            print("✗ Error: Tidak dapat mendeteksi jumlah baris")
            return
        
        processed_count = 0
        updated_count = 0
        
        print(f"\n{'='*60}")
        print(f"Memproses {total_rows - start_row + 1} baris data...")
        print(f"{'='*60}\n")
        
        # Reset tracking variables
        self.previous_docket = None
        self.previous_beban = None
        
        for row_index in range(start_row, total_rows + 1):
            processed_count += 1
            
            # Hitung BEBAN
            beban_value = self.calculate_beban(row_index)
            
            if beban_value is not None:
                # Update nilai BEBAN di Excel
                self.set_cell_value(row_index, 'BEBAN', beban_value)
                updated_count += 1
                
                # Log progress setiap 10 baris
                if processed_count % 10 == 0:
                    print(f"  Progress: {processed_count}/{total_rows - start_row + 1} baris | "
                          f"Terupdate: {updated_count}")
        
        print(f"\n{'='*60}")
        print(f"✓ Proses selesai!")
        print(f"  Total baris diproses: {processed_count}")
        print(f"  Total BEBAN terupdate: {updated_count}")
        print(f"{'='*60}\n")
    
    def save_excel(self) -> bool:
        """Menyimpan perubahan ke file Excel yang sama"""
        if self.workbook is None:
            print("✗ Error: Workbook belum dimuat")
            return False
        
        try:
            self.workbook.save(self.file_path)
            print(f"✓ File berhasil disimpan: {self.file_path}")
            return True
        except Exception as e:
            print(f"✗ Error saat menyimpan file: {str(e)}")
            return False
    
    def close(self) -> None:
        """Menutup workbook"""
        if self.workbook:
            self.workbook.close()
            print("✓ Workbook ditutup")


def main() -> None:
    """Fungsi utama untuk menjalankan processor"""
    
    # ========== KONFIGURASI ==========
    # Ganti dengan path file Excel Anda
    script_dir = Path(__file__).parent
    file_path = script_dir / "PENGUJIAN.xlsx"
    SHEET_NAME = "ODOO"  # Nama sheet yang akan diproses
    START_ROW = 2  # Baris mulai (2 = skip header)
    # =================================
    
    print("\n" + "="*60)
    print("  PROCESSOR PERHITUNGAN BEBAN - FILE EXCEL")
    print("="*60 + "\n")
    
    # Inisialisasi processor
    processor = ExcelBebanProcessor(file_path, SHEET_NAME) # type: ignore
    
    # Load file Excel
    if not processor.load_excel():
        return
    
    # Proses semua baris
    processor.process_all_rows(start_row=START_ROW)
    
    # Simpan hasil
    processor.save_excel()
    
    # Tutup workbook
    processor.close()
    
    print("\n✓ Proses selesai dengan sukses!\n")


if __name__ == "__main__":
    main()