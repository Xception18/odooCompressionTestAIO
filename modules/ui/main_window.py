import tkinter as tk
import traceback
import os
import threading
import logging
import queue
import customtkinter as ctk
import sys
import time
from pathlib import Path
from tkinter import ttk, messagebox
from tkinter.scrolledtext import ScrolledText

from modules.utils import resource_path, ThreadSafeLogHandler

import pandas as pd
import openpyxl
import requests
import glob
from datetime import datetime
import configparser
import xlwings as xw
from sqlalchemy import create_engine
from tkinter import filedialog

# Import other necessary modules
# Note: Assuming these are still in the root or moved appropriately
try:
    from modules.daemon_sync import threadSinkData, setup_daemon_logging
    from modules.excel_handler import ExcelBebanProcessor
except ImportError:
    # Fallback if running from modules/ui
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
    from modules.daemon_sync import threadSinkData, setup_daemon_logging
    from modules.excel_handler import ExcelBebanProcessor

class ExcelProcessorGUI:
    def __init__(self, root):
        self.root = root
        # Ensure logger is always a logging.Logger instance to avoid None attribute errors
        # It will be reconfigured properly in setup_logging()
        self.logger = logging.getLogger('ExcelProcessor')
        self.daemon = None
        self.daemon_running = False
        # If no handlers are attached yet, attach a NullHandler to avoid "No handler found" warnings
        if not self.logger.handlers:
            self.logger.addHandler(logging.NullHandler())
        self.root.title("Excel Data Processor - ODOO Test")
        self.root.geometry("1100x800")
        # Add process tracking for Rencana Benda Uji
        self.rencana_process_running = False
        self.rencana_stop_event = threading.Event()
        self.rencana_thread = None
        # Add thread-safe logging queue
        self.log_queue = queue.Queue()
        # Start queue processor
        self.process_log_queue()
        # Set window close protocol
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        # configure the root window
        self.root.configure(bg='#212121')
        
        # Variables
        self.file_excel_1 = tk.StringVar()
        self.file_excel_2 = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.file_csv = tk.StringVar()
        self.db_user = tk.StringVar(value='postgres')
        self.db_password = tk.StringVar(value='adhimix')
        self.db_host = tk.StringVar(value='localhost')
        self.db_port = tk.StringVar(value='5432')
        self.db_name = tk.StringVar(value='alatujidb')
        self.table_name = tk.StringVar(value='pengujian')
        
        # Default values
        # We need to adjust paths since we are in modules/ui
        # But __file__ will be modules/ui/main_window.py
        # The original code expected to be in root.
        # Let's use resource_path or relative to root.
        
        # Assuming the app is run from root, we can use os.getcwd() or relative paths
        # But better to rely on the structure.
        
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
        
        excel_1 = os.path.join(base_dir, 'Rencana', 'RENCANA.xlsx')
        excel_2 = os.path.join(base_dir, 'Pengujian', 'PENGUJIAN.xlsx')
        output = os.path.join(base_dir, 'Output')
        
        self.file_excel_1.set(str(excel_1))
        self.file_excel_2.set(str(excel_2))
        self.output_dir.set(str(output))
        self.setup_ui()
        self.setup_logging()


    def setup_ui(self):
        # Main title
        title_label = ctk.CTkLabel(
            self.root, 
            text="Excel Data Processor", 
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="#4A90E2"
        )
        title_label.pack(pady=(20, 10))
        
        subtitle_label = ctk.CTkLabel(
            self.root, 
            text="ODOO Test Integration", 
            font=ctk.CTkFont(size=16),
            text_color="#B0B0B0"
        )
        subtitle_label.pack(pady=(0, 20))
        
        # Main tabview
        self.tabview = ctk.CTkTabview(self.root, width=1000, height=600)
        self.tabview.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Add tabs
        self.setup_file_configure_tab()
        self.setup_db_configure_tab()
        self.setup_process_tab()
        self.setup_rencana_benda_uji_tab()
        self.setup_log_tab()
        
    def setup_file_configure_tab(self):
        # File Configuration Tab
        file_tab = self.tabview.add("📁 File Configuration")
        
        # Create scrollable frame
        scrollable_frame = ctk.CTkScrollableFrame(file_tab, width=950, height=500)
        scrollable_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # File selection section
        file_section = ctk.CTkFrame(scrollable_frame)
        file_section.pack(fill="x", pady=(0, 20))
        
        section_title = ctk.CTkLabel(
            file_section, 
            text="📂 File Selection", 
            font=ctk.CTkFont(size=18, weight="bold")
        )
        section_title.pack(pady=(15, 10))
        
        # Source Excel File
        self.create_file_input(file_section, "Source Excel File (Rencana):", self.file_excel_1, 0)
        
        # Target Excel File
        self.create_file_input(file_section, "Target Excel File (Pengujian):", self.file_excel_2, 1)
        
        # Output Directory
        self.create_directory_input(file_section, "Output Directory:", self.output_dir, 2)
        
        # Sheet Configuration Section
        sheet_section = ctk.CTkFrame(scrollable_frame)
        sheet_section.pack(fill="x", pady=(0, 20))
        
        sheet_title = ctk.CTkLabel(
            sheet_section, 
            text="📋 Sheet Configuration", 
            font=ctk.CTkFont(size=18, weight="bold")
        )
        sheet_title.pack(pady=(15, 10))
        
        # Create a grid frame for sheet settings
        grid_frame = ctk.CTkFrame(sheet_section)
        grid_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # configure grid weights
        grid_frame.grid_columnconfigure(1, weight=1)
        grid_frame.grid_columnconfigure(3, weight=1)
        
        # Sheet name variables
        self.nama_sheet_1 = tk.StringVar(value='Report Excel')
        self.nama_sheet_2 = tk.StringVar(value='ODOO')
        self.nama_sheet_3 = tk.StringVar(value='pengujian')
        self.nama_sheet_4 = tk.StringVar(value='bjdt_id')
        
        # Create sheet input fields
        self.create_sheet_input(grid_frame, "Source Sheet Name:", self.nama_sheet_1, 0, 0)
        self.create_sheet_input(grid_frame, "Target Sheet ODOO:", self.nama_sheet_2, 0, 2)
        self.create_sheet_input(grid_frame, "Sheet Pengujian:", self.nama_sheet_3, 1, 0)
        self.create_sheet_input(grid_frame, "Sheet BJDT ID:", self.nama_sheet_4, 1, 2)
        
    def create_file_input(self, parent, label_text, variable, row):
        frame = ctk.CTkFrame(parent)
        frame.pack(fill="x", padx=20, pady=5)
        
        label = ctk.CTkLabel(frame, text=label_text, font=ctk.CTkFont(size=14))
        label.pack(anchor="w", padx=15, pady=(10, 5))
        
        input_frame = ctk.CTkFrame(frame)
        input_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        entry = ctk.CTkEntry(input_frame, textvariable=variable, height=35, font=ctk.CTkFont(size=12))
        entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        browse_btn = ctk.CTkButton(
            input_frame, 
            text="Browse", 
            command=lambda: self.browse_file(variable),
            width=100,
            height=35
        )
        browse_btn.pack(side="right")
        
    def create_directory_input(self, parent, label_text, variable, row):
        frame = ctk.CTkFrame(parent)
        frame.pack(fill="x", padx=20, pady=5)
        
        label = ctk.CTkLabel(frame, text=label_text, font=ctk.CTkFont(size=14))
        label.pack(anchor="w", padx=15, pady=(10, 5))
        
        input_frame = ctk.CTkFrame(frame)
        input_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        entry = ctk.CTkEntry(input_frame, textvariable=variable, height=35, font=ctk.CTkFont(size=12))
        entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        browse_btn = ctk.CTkButton(
            input_frame, 
            text="Browse", 
            command=self.browse_directory,
            width=100,
            height=35
        )
        browse_btn.pack(side="right")
        
    def create_sheet_input(self, parent, label_text, variable, row, col):
        label = ctk.CTkLabel(parent, text=label_text, font=ctk.CTkFont(size=12))
        label.grid(row=row*2, column=col, sticky="w", padx=15, pady=(10, 5))
        
        entry = ctk.CTkEntry(parent, textvariable=variable, height=30, font=ctk.CTkFont(size=11))
        entry.grid(row=row*2+1, column=col, sticky="ew", padx=15, pady=(0, 15))
        
    def setup_db_configure_tab(self):
        # Database Configuration Tab
        db_tab = self.tabview.add("🗄️ Database configure")
        
        # Create main frame
        main_frame = ctk.CTkFrame(db_tab)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(
            main_frame, 
            text="🗄️ Database Configuration", 
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title.pack(pady=(20, 30))
        
        # Create grid frame
        grid_frame = ctk.CTkFrame(main_frame)
        grid_frame.pack(pady=20)
        
        # Database input fields
        self.create_db_input(grid_frame, "Database User:", self.db_user, 0, show_password=False)
        self.create_db_input(grid_frame, "Database Password:", self.db_password, 1, show_password=True)
        self.create_db_input(grid_frame, "Database Host:", self.db_host, 2, show_password=False)
        self.create_db_input(grid_frame, "Database Port:", self.db_port, 3, show_password=False)
        self.create_db_input(grid_frame, "Database Name:", self.db_name, 4, show_password=False)
        self.create_db_input(grid_frame, "Table Name:", self.table_name, 5, show_password=False)
        
        # Test connection button
        test_btn = ctk.CTkButton(
            main_frame, 
            text="🔍 Test Connection", 
            command=self.test_db_connection,
            width=200,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        test_btn.pack(pady=30)
        
    def create_db_input(self, parent, label_text, variable, row, show_password=False):
        label = ctk.CTkLabel(parent, text=label_text, font=ctk.CTkFont(size=14))
        label.grid(row=row, column=0, sticky="w", padx=20, pady=10)
        
        if show_password:
            entry = ctk.CTkEntry(parent, textvariable=variable, show='*', width=300, height=35, font=ctk.CTkFont(size=12))
        else:
            entry = ctk.CTkEntry(parent, textvariable=variable, width=300, height=35, font=ctk.CTkFont(size=12))
        entry.grid(row=row, column=1, padx=20, pady=10)
        
    def setup_process_tab(self):
        # Process Control Tab
        process_tab = self.tabview.add("⚙️ Process Control")
        
        # Main frame
        main_frame = ctk.CTkFrame(process_tab)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(
            main_frame, 
            text="⚙️ Process Control Center", 
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title.pack(pady=(20, 30))
        
        # Progress section
        progress_frame = ctk.CTkFrame(main_frame)
        progress_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        progress_label = ctk.CTkLabel(progress_frame, text="📊 Progress Status", font=ctk.CTkFont(size=16, weight="bold"))
        progress_label.pack(pady=(15, 10))
        
        # Progress bar (using tkinter since CTk doesn't have indeterminate mode)
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate', style="Custom.Horizontal.TProgressbar")
        self.progress.pack(fill='x', padx=20, pady=10)
        
        # Status label
        self.status_label = ctk.CTkLabel(progress_frame, text="Ready to process", font=ctk.CTkFont(size=14))
        self.status_label.pack(pady=(0, 15))
        
        # Buttons section
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", padx=20, pady=20)
        
        button_title = ctk.CTkLabel(button_frame, text="🚀 Available Actions", font=ctk.CTkFont(size=16, weight="bold"))
        button_title.pack(pady=(15, 20))
        
        # First row of buttons
        button_row1 = ctk.CTkFrame(button_frame)
        button_row1.pack(pady=5)
        
        self.process_btn = ctk.CTkButton(
            button_row1, 
            text="🚀 Start Full Process", 
            command=self.start_process,
            width=180,
            height=40,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.process_btn.pack(side='left', padx=5)
        
        self.copy_btn = ctk.CTkButton(
            button_row1, 
            text="📋 Copy Data", 
            command=self.copy_data_only,
            width=180,
            height=40,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.copy_btn.pack(side='left', padx=5)
        
        self.api_btn = ctk.CTkButton(
            button_row1, 
            text="🔍 Get BJDT ID", 
            command=self.get_bjdt_id_only,
            width=180,
            height=40,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.api_btn.pack(side='left', padx=5)
        
        # Second row of buttons
        button_row2 = ctk.CTkFrame(button_frame)
        button_row2.pack(pady=5)
        
        self.csv_btn = ctk.CTkButton(
            button_row2, 
            text="📄 Generate CSV", 
            command=self.generate_csv_files,
            width=180,
            height=40,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.csv_btn.pack(side='left', padx=5)
        
        self.upload_csv_btn = ctk.CTkButton(
            button_row2, 
            text="📤 Upload CSV File", 
            command=lambda: self.browse_files(self.file_csv),
            width=180,
            height=40,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.upload_csv_btn.pack(side='left', padx=5)
        
        self.upload_db_btn = ctk.CTkButton(
            button_row2, 
            text="🗄️ Upload to DB", 
            command=self.upload_to_db_only,
            width=180,
            height=40,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.upload_db_btn.pack(side='left', padx=5)


        button_row3 = ctk.CTkFrame(button_frame)
        button_row3.pack(pady=5)
        
        self.grid_btn = ctk.CTkButton(
            button_row3, 
            text="🧾 Open Grid Benda Uji", 
            command=self.launch_grid_benda_uji,
            width=180,
            height=40,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.grid_btn.pack(side='left', padx=5)
        
        # Sync button (separate)
        sync_frame = ctk.CTkFrame(main_frame)
        sync_frame.pack(fill="x", padx=10, pady=10)
        
        self.sync_btn = ctk.CTkButton(
            sync_frame, 
            text="🔄 Continue to ODOO Sync", 
            command=self.toggle_sync,
            width=250,
            height=45,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="#28A745",
            hover_color="#218838"
        )
        self.sync_btn.pack(pady=20)
        
    def setup_log_tab(self):
        # Logs Tab
        log_tab = self.tabview.add("📝 Process Logs")
        
        # Main frame
        main_frame = ctk.CTkFrame(log_tab)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(
            main_frame, 
            text="📝 Process Logs", 
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title.pack(pady=(15, 10))
        
        # Log frame
        log_frame = ctk.CTkFrame(main_frame)
        log_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Log text widget (using tkinter ScrolledText for better functionality)
        self.log_text = ScrolledText(
            log_frame, 
            height=25, 
            width=100,
            bg='#2B2B2B',
            fg='#FFFFFF',
            font=('Consolas', 10),
            insertbackground='white',
            selectbackground='#4A4A4A'
        )
        self.log_text.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Clear button
        clear_btn = ctk.CTkButton(
            main_frame, 
            text="🗑️ Clear Logs", 
            command=self.clear_logs,
            width=150,
            height=35,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#DC3545",
            hover_color="#C82333"
        )
        clear_btn.pack(pady=(0, 15))

    def setup_rencana_benda_uji_tab(self):
        """Setup tab for Rencana Benda Uji input"""
        grid_tab = self.tabview.add("🧾 Input Rencana Benda Uji")

        frame = ctk.CTkFrame(grid_tab)
        frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Title
        title = ctk.CTkLabel(
            frame, 
            text="Input Rencana Benda Uji", 
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title.pack(pady=(20, 10))

        # Description
        desc = ctk.CTkLabel(
            frame, 
            text="Process input data using Selenium automation\n"
                "Click 'Start' to begin, 'Stop' to terminate process", 
            font=ctk.CTkFont(size=12),
            text_color="#B0B0B0"
        )
        desc.pack(pady=(0, 20))

        # Launch button (toggle start/stop)
        self.rencana_btn = ctk.CTkButton(
            frame,
            text="🚀 Start Process",
            width=220,
            height=45,
            command=self.launch_rencana_benda_uji,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.rencana_btn.pack(pady=(20, 10))

        # Status label
        self.rencana_status = ctk.CTkLabel(
            frame, 
            text="Ready to process", 
            font=ctk.CTkFont(size=13),
            text_color="#4A90E2"
        )
        self.rencana_status.pack(pady=(10, 20))
        
        # Info frame
        info_frame = ctk.CTkFrame(frame)
        info_frame.pack(fill='x', padx=20, pady=(20, 0))
        
        info_title = ctk.CTkLabel(
            info_frame,
            text="ℹ️ Information",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        info_title.pack(pady=(10, 5))
        
        info_text = ctk.CTkLabel(
            info_frame,
            text="• Process runs in background\n"
                "• GUI remains responsive during processing\n"
                "• Click Stop button to terminate safely\n"
                "• Selenium browser will close automatically",
            font=ctk.CTkFont(size=11),
            justify='left'
        )
        info_text.pack(pady=(0, 10), padx=15)
            
    def launch_rencana_benda_uji(self):
        """Toggle Rencana Benda Uji process - Start or Stop"""
        if self.rencana_process_running:
            # Stop the process
            self.stop_rencana_process()
        else:
            # Start the process
            self.start_rencana_process()


    def start_rencana_process(self):
        """Start Rencana Benda Uji process in background thread"""
        # Clear stop event
        self.rencana_stop_event.clear()
        
        def run_process():
            selenium_driver = None
            try:
                # Import here to avoid circular dependency if possible, or just standard import
                # Assuming input_rencana_benda_uji is in root
                sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
                from modules.input_rencana_benda_uji import run_with_custom_path_and_stop
                from pathlib import Path
                
                # Update UI
                def update_ui_start():
                    if hasattr(self, 'rencana_status'):
                        self.rencana_status.configure(text="🔄 Processing...")
                    if hasattr(self, 'rencana_btn'):
                        self.rencana_btn.configure(
                            text="⏹️ Stop Process",
                            fg_color="#DC3545",
                            hover_color="#C82333"
                        )
                    if hasattr(self, 'progress'):
                        self.progress.start()
                
                self.root.after(0, update_ui_start)
                self.rencana_process_running = True
                
                # Get file path
                # Use the base_dir calculated in __init__
                base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
                excel_file_path = os.path.join(base_dir, 'Data Input Pengujian', 'data.xlsx')
                
                if self.logger:
                    self.logger.info("=" * 60)
                    self.logger.info("Starting Rencana Benda Uji process...")
                    self.logger.info(f"Excel file: {excel_file_path}")
                    self.logger.info("=" * 60)
                
                # Modified run to support cancellation
                results = self.run_with_stop_event(excel_file_path)
                
                # Handle completion
                def on_complete():
                    self.rencana_process_running = False
                    
                    if hasattr(self, 'progress'):
                        self.progress.stop()
                    
                    if self.rencana_stop_event.is_set():
                        # Process was stopped by user
                        status_msg = "Process stopped by user"
                        if hasattr(self, 'rencana_status'):
                            self.rencana_status.configure(text=status_msg)
                        if self.logger:
                            self.logger.warning(status_msg)
                        messagebox.showinfo("Stopped", "Proses dihentikan oleh user")
                        
                    elif results:
                        # Process completed successfully
                        success_msg = f"✅ Completed: {results['successful_rows']} rows"
                        if hasattr(self, 'rencana_status'):
                            self.rencana_status.configure(text=success_msg)
                        
                        if self.logger:
                            self.logger.info("=" * 60)
                            self.logger.info(f"Process completed successfully")
                            self.logger.info(f"Successful rows: {results['successful_rows']}")
                            self.logger.info(f"Failed rows: {results.get('failed_rows', 0)}")
                            self.logger.info("=" * 60)
                        
                        messagebox.showinfo("Success", 
                            f"Proses selesai!\n\n"
                            f"✅ Berhasil: {results['successful_rows']} baris\n"
                            f"❌ Gagal: {results.get('failed_rows', 0)} baris")
                    else:
                        # Process failed
                        error_msg = "❌ Process failed"
                        if hasattr(self, 'rencana_status'):
                            self.rencana_status.configure(text=error_msg)
                        
                        if self.logger:
                            self.logger.error("Process returned no results")
                        
                        messagebox.showerror("Error", "Proses gagal. Periksa log untuk detail.")
                    
                    # Reset button
                    if hasattr(self, 'rencana_btn'):
                        self.rencana_btn.configure(
                            text="🚀 Start Process",
                            fg_color=["#3B8ED0", "#1F6AA5"],
                            hover_color=["#36719F", "#144870"]
                        )
                
                self.root.after(0, on_complete)
                
            except Exception as e:
                def on_error():
                    self.rencana_process_running = False
                    error_msg = f"❌ Error: {str(e)}"
                    
                    if hasattr(self, 'rencana_status'):
                        self.rencana_status.configure(text=error_msg)
                    if hasattr(self, 'rencana_btn'):
                        self.rencana_btn.configure(
                            text="🚀 Start Process",
                            fg_color=["#3B8ED0", "#1F6AA5"],
                            hover_color=["#36719F", "#144870"]
                        )
                    if hasattr(self, 'progress'):
                        self.progress.stop()
                    
                    if self.logger:
                        self.logger.error(f"Exception in Rencana Benda Uji: {str(e)}")
                        self.logger.error(traceback.format_exc())
                    
                    messagebox.showerror("Error", f"Terjadi kesalahan:\n{str(e)}")
                
                self.root.after(0, on_error)
            finally:
                # Cleanup
                self.rencana_process_running = False
                self.rencana_thread = None
        
        # Start in background thread
        self.rencana_thread = threading.Thread(
            target=run_process, 
            daemon=True, 
            name="RencanaBendaUjiThread"
        )
        self.rencana_thread.start()
        
        if self.logger:
            self.logger.info("Rencana Benda Uji process started in background thread")

    def stop_rencana_process(self):
        """Stop the running Rencana Benda Uji process"""
        if not self.rencana_process_running:
            return
        
        if self.logger:
            self.logger.warning("=" * 60)
            self.logger.warning("STOP signal received - Stopping Rencana Benda Uji process...")
            self.logger.warning("=" * 60)
            # Driver cleanup is handled by the background thread when it sees the stop event

        
        # Set stop event
        self.rencana_stop_event.set()
        
        # Update UI immediately
        if hasattr(self, 'rencana_status'):
            self.rencana_status.configure(text="⏹️ Stopping process...")
        
        if hasattr(self, 'rencana_btn'):
            self.rencana_btn.configure(state='disabled', text="⏹️ Stopping...")
        
        # Wait for thread to finish (with timeout)
        if self.rencana_thread and self.rencana_thread.is_alive():
            self.rencana_thread.join(timeout=1.0)
            
            if self.rencana_thread.is_alive():
                if self.logger:
                    self.logger.warning("Process did not stop gracefully within timeout")
            else:
                if self.logger:
                    self.logger.info("Process stopped successfully")
        
        # Reset state
        self.rencana_process_running = False
        
        # Re-enable button
        if hasattr(self, 'rencana_btn'):
            self.rencana_btn.configure(
                state='normal',
                text="🚀 Start Process",
                fg_color=["#3B8ED0", "#1F6AA5"],
                hover_color=["#36719F", "#144870"]
            )

    def run_with_stop_event(self, excel_file_path):
        """
        Wrapper untuk run_with_custom_path dengan stop event support.
        Ini memerlukan modifikasi pada input_rencana_benda_uji.py
        """
        try:
            sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
            from modules.input_rencana_benda_uji import run_with_custom_path_and_stop
            
            # Jika fungsi dengan stop event tersedia
            return run_with_custom_path_and_stop(excel_file_path, self.rencana_stop_event)
        except ImportError:
            # Fallback ke fungsi original jika tidak ada stop event support
            from modules.input_rencana_benda_uji import run_with_custom_path_and_stop
            
            if self.logger:
                self.logger.warning("Stop event not supported in input_rencana_benda_uji module")
                self.logger.warning("Using original function - manual stop may not work properly")
            
            return run_with_custom_path_and_stop(excel_file_path)
        
    def launch_grid_benda_uji(self):
        """Launch the Grid Benda Uji as a separate process to avoid wx/tkinter threading conflicts."""
        try:
            import subprocess
            
            # Get the path to the grid_benda_uji.py script
            # We need to be careful about paths. 
            # If running from source:
            base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
            script_path = os.path.join(base_dir, 'modules', 'grid_benda_uji.py')
            
            if not os.path.exists(script_path):
                # Try alternative location if packaged or structure is different
                script_path = os.path.join(os.getcwd(), 'modules', 'grid_benda_uji.py')
                
            if not os.path.exists(script_path):
                messagebox.showerror("Error", f"Could not find grid_benda_uji.py at:\n{script_path}")
                return

            if self.logger:
                self.logger.info(f"Launching Grid Benda Uji from: {script_path}")

            # Launch as separate process
            # Use sys.executable to ensure we use the same python interpreter
            subprocess.Popen([sys.executable, script_path], cwd=base_dir)
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"Failed to launch Grid Benda Uji: {e}")
                self.logger.error(traceback.format_exc())
            messagebox.showerror("Error", f"Failed to launch Grid Benda Uji:\n{str(e)}")

    def setup_logging(self):
        """Setup logging configuration"""
        # Configure root logger
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler("aplikasiAlatUji.log"),
                logging.StreamHandler(sys.stdout)
            ]
        )
        
        # Configure application logger
        self.logger = logging.getLogger('ExcelProcessor')
        self.logger.setLevel(logging.INFO)
        
        # Add thread-safe handler
        queue_handler = ThreadSafeLogHandler(self.log_queue)
        self.logger.addHandler(queue_handler)

    def process_log_queue(self):
        """Process messages from the log queue and update GUI"""
        try:
            while True:
                record = self.log_queue.get_nowait()
                if hasattr(self, 'log_text'):
                    self.log_text.insert('end', record + '\n')
                    self.log_text.see('end')
        except queue.Empty:
            pass
        finally:
            # Schedule next check
            self.root.after(100, self.process_log_queue)

    def clear_logs(self):
        """Clear the log text widget"""
        if hasattr(self, 'log_text'):
            self.log_text.delete('1.0', 'end')
            self.logger.info("Logs cleared")

    def browse_file(self, variable):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if filename:
            variable.set(filename)

    def browse_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_dir.set(directory)

    def browse_files(self, variable):
        filename = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv")]
        )
        if filename:
            variable.set(filename)

    def on_closing(self):
        """Handle window closing"""
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.daemon_running = False
            self.root.destroy()
            sys.exit(0)

    # Placeholder methods for buttons not fully implemented in this snippet
    # You would need to move the actual logic for these methods as well
    def update_status(self, message):
        """Update the status label safely from any thread"""
        def _update():
            if hasattr(self, 'status_label'):
                self.status_label.configure(text=message)
        self.root.after(0, _update)

    def disable_buttons(self):
        self.process_btn.configure(state='disabled')
        self.copy_btn.configure(state='disabled')
        self.api_btn.configure(state='disabled')
        self.csv_btn.configure(state='disabled')
        self.upload_csv_btn.configure(state='disabled')
        self.upload_db_btn.configure(state='disabled')
        self.progress.start()
        
    def enable_buttons(self):
        self.process_btn.configure(state='normal')
        self.copy_btn.configure(state='normal')
        self.api_btn.configure(state='normal')
        self.sync_btn.configure(state='normal')
        self.csv_btn.configure(state='normal')
        self.upload_csv_btn.configure(state='normal')
        self.upload_db_btn.configure(state='normal')
        self.progress.stop()

    def start_process(self):
        threading.Thread(target=self.full_process, daemon=True).start()

    def full_process(self):
        self.disable_buttons()
        try:
            self.update_status("Starting full process...")
            
            # Step 1: Copy data
            if self.logger is None:
                self.setup_logging()
            self.logger.info("=== Starting Data Copy Process ===")
            self.copy_excel_data()
            
            # Step 2: Get BJDT ID
            self.logger.info("=== Starting BJDT ID Process ===")
            self.get_bjdt_ids()
            
            # Step 3: Generate CSV files
            self.logger.info("=== Starting CSV Generation ===")
            self.generate_csv_files()
            messagebox.showinfo("Success", "Generate CSV files separated by date successfully!")
            
            # Step 4: Upload to database
            self.logger.info("=== Starting Database Upload ===")
            self.upload_to_database()
            
            self.update_status("Full process completed successfully!")
            self.sync_btn.configure(state='normal')
            messagebox.showinfo("Success", "Full process completed successfully!")
            
        except Exception as e:
            self.logger.error(f"Error in full process: {str(e)}")
            self.logger.error(traceback.format_exc())
            self.update_status("Process failed!")
            messagebox.showerror("Error", f"Process failed: {str(e)}")
        finally:
            self.enable_buttons()

    def copy_data_only(self):
        threading.Thread(target=self.copy_data_process, daemon=True).start()

    def copy_data_process(self):
        self.disable_buttons()
        try:
            self.update_status("Copying data...")
            self.copy_excel_data()
            self.update_status("Data copy completed!")
            messagebox.showinfo("Success", "Data copied successfully!")
        except Exception as e:
            self.logger.error(f"Error copying data: {str(e)}")
            self.update_status("Copy failed!")
            messagebox.showerror("Error", f"Copy failed: {str(e)}")
        finally:
            self.enable_buttons()

    def get_bjdt_id_only(self):
        threading.Thread(target=self.bjdt_id_process, daemon=True).start()

    def bjdt_id_process(self):
        self.disable_buttons()
        try:
            self.update_status("Getting BJDT IDs...")
            self.get_bjdt_ids()
            self.update_status("BJDT ID process completed!")
            messagebox.showinfo("Success", "BJDT IDs retrieved successfully!")
        except Exception as e:
            self.logger.error(f"Error getting BJDT IDs: {str(e)}")
            self.update_status("BJDT ID process failed!")
            messagebox.showerror("Error", f"BJDT ID process failed: {str(e)}")
        finally:
            self.enable_buttons()

    def generate_csv_files(self):
        # This calls the actual logic which we will implement next
        # But for the button command, we might want to run it in thread if it takes long
        # The original code had generate_csv_files doing the work directly or via thread?
        # Original code: command=self.generate_csv_files
        # And generate_csv_files did the work.
        # Let's keep it simple and run in thread to avoid freezing UI
        threading.Thread(target=self._generate_csv_files_thread, daemon=True).start()

    def _generate_csv_files_thread(self):
        self.disable_buttons()
        try:
            self.update_status("Generating CSV files...")
            self._generate_csv_files_logic()
            self.update_status("CSV generation completed!")
        except Exception as e:
            self.logger.error(f"Error generating CSV: {str(e)}")
            self.update_status("CSV generation failed!")
            messagebox.showerror("Error", f"CSV generation failed: {str(e)}")
        finally:
            self.enable_buttons()

    def upload_to_db_only(self):
        threading.Thread(target=self.upload_db_process, daemon=True).start()

    def upload_db_process(self):
        self.disable_buttons()
        try:
            self.update_status("Uploading to database...")
            self.upload_to_database_files()
        except Exception as e:
            self.logger.error(f"Error uploading to database: {str(e)}")
            self.update_status("Database upload failed!")
            messagebox.showerror("Error", f"Database upload failed: {str(e)}")
        finally:
            self.enable_buttons()
            self.update_status("Database upload completed!")
            messagebox.showinfo("Success", "Data uploaded to database successfully!")
            
    def copy_excel_data(self):
        """Copy data from source Excel to target Excel"""
        try:
        # Column names to copy
            kolom_yang_dicopy = ['Docket', 'Nomor Kontrak', 'Nomor SPP', 'Proyek', 'Kontraktor', 'Mutu', 'No. Urut', 
                            'Tanggal Rencana', 'Rencana Umur Test', 'Bentuk Benda Uji', 'Target (%)', 
                            'Tanggal Realisasi Test', 'Realisasi Umur Test', 'Hasil Test', 'Rusak', 
                            'Kode Benda Uji', 'Jenis Retakan', 'Kn', 'Mpa', 'Kg/cm2', 'Berat (kg)', 
                            'Persentase Kekuatan (%)']
            
            # Load source workbook
            self.logger.info("Loading source Excel file...")
            wb1 = openpyxl.load_workbook(self.file_excel_1.get())
            sheet1 = wb1[self.nama_sheet_1.get()]
            
            # Get header and column indices
            header = [cell.value for cell in sheet1[1]]
            index_kolom = [header.index(k) + 1 for k in kolom_yang_dicopy if k in header]
            
            self.logger.info(f"Found {len(index_kolom)} columns to copy")
            
            # Extract data from specified columns
            data_kolom = []
            for col in index_kolom:
                col_data = []
                for row in sheet1.iter_rows(min_row=2, min_col=col, max_col=col):
                    val = row[0].value
                    col_data.append(val)
                data_kolom.append(col_data)
            
            # Check for duplicates (simplified version)
            total_rows = len(data_kolom[0]) if data_kolom else 0
            self.logger.info(f"Total rows to process: {total_rows}")
            
            # Load target workbook
            self.logger.info("Loading target Excel file...")
            wb2 = openpyxl.load_workbook(self.file_excel_2.get())
            sheet2 = wb2[self.nama_sheet_2.get()]
            
            # Write data to target sheet
            baris_mulai = 2
            kolom_mulai = 1
            
            self.logger.info("Copying data to target Excel...")
            for i, kolom_data in enumerate(data_kolom):
                for j, value in enumerate(kolom_data):
                    sheet2.cell(row=baris_mulai + j, column=kolom_mulai + i, value=value)
            
            # Save target workbook before processing BEBAN
            wb2.save(self.file_excel_2.get())
            wb2.close()
            # Load the Excel file
            beban_processor = ExcelBebanProcessor(self.file_excel_2.get(), self.nama_sheet_2.get())
            if beban_processor.load_excel():
                # Process all rows to calculate BEBAN
                # Start from row 2 (assuming row 1 is header)
                beban_processor.process_all_rows(start_row=2)
                
                # Save the results
                if beban_processor.save_excel():
                    self.logger.info("BEBAN values calculated and saved successfully")
                else:
                    self.logger.error("Failed to save BEBAN calculations")
                
                # Close the workbook
                beban_processor.close()
            else:
                self.logger.error("Failed to load Excel file for BEBAN processing")
            
            # ============================================================
            # END OF BEBAN CALCULATION
            # ============================================================
            
            # Update idpengujian in pengujian sheet
            self.logger.info("Updating idpengujian...")
            self.update_idpengujian()
            
            self.logger.info("All processes completed successfully!")
            
        except FileNotFoundError as e:
            self.logger.error(f"File not found: {str(e)}")
            raise
        except KeyError as e:
            self.logger.error(f"Sheet not found: {str(e)}")
            raise
        except Exception as e:
            self.logger.error(f"Error during copy_excel_data: {str(e)}")
            raise
        
    def update_idpengujian(self):
        """Update idpengujian in pengujian sheet"""
        csv_files = glob.glob(os.path.join(self.output_dir.get(), "*.csv"))
        if csv_files:
            latest_csv = max(csv_files, key=os.path.getctime)
            df_latest = pd.read_csv(latest_csv, sep=';', encoding='utf-8-sig')
            if 'idpengujian' in df_latest.columns and not df_latest['idpengujian'].empty:
                last_id = pd.to_numeric(df_latest['idpengujian'], errors='coerce').dropna().astype(int).max()
                next_id = last_id + 1
            else:
                next_id = 1
        else:
            next_id = 1
        
        # Update idpengujian value in pengujian sheet
        wb_pengujian = openpyxl.load_workbook(self.file_excel_2.get())
        ws_pengujian = wb_pengujian[self.nama_sheet_3.get()]
        
        # Find idpengujian column
        for col in range(1, ws_pengujian.max_column + 1):
            if ws_pengujian.cell(row=1, column=col).value == 'idpengujian':
                # openpyxl may return MergedCell objects for some sheets; assignment is intended at runtime
                ws_pengujian.cell(row=2, column=col).value = next_id  # type: ignore
                break
        
        wb_pengujian.save(self.file_excel_2.get())
        wb_pengujian.close()
        self.logger.info(f"Updated idpengujian to: {next_id}")
        
    def get_bjdt_ids(self):
        """Get BJDT IDs from API"""
        # Calculate formulas first
        with xw.App(visible=False) as app:
            wb = app.books.open(self.file_excel_2.get())
            wb.app.calculate()
            wb.save()
            wb.close()
        
        # Load workbooks
        wb_data = openpyxl.load_workbook(self.file_excel_2.get(), data_only=True)
        sheet_data = wb_data[self.nama_sheet_4.get()]
        
        wb_write = openpyxl.load_workbook(self.file_excel_2.get())
        sheet_write = wb_write[self.nama_sheet_4.get()]
        
        # Get header and column indices
        header = [cell.value for cell in sheet_data[1]]
        
        try:
            index_doc_no = header.index('Docket')
            index_no_urut = header.index('No URUT')
        except ValueError as e:
            self.logger.error(f"Required column not found: {e}")
            return
        
        # Add ID column if not exists
        if 'ID' not in header:
            sheet_write.cell(row=1, column=len(header) + 1, value='ID')
            index_bjdt_id = len(header)
        else:
            index_bjdt_id = header.index('ID')
        
        error_logs = []
        
        # Process each row
        for i, (row_data, row_write) in enumerate(zip(
            sheet_data.iter_rows(min_row=2), 
            sheet_write.iter_rows(min_row=2)
        ), start=2):
            
            doc_no = row_data[index_doc_no].value
            no_urut = row_data[index_no_urut].value
            
            if doc_no and no_urut:
                doc_no = str(doc_no).strip()
                no_urut = str(no_urut).strip()
                url = f"https://rmc.adhimix.web.id/benda_uji/?doc_no={doc_no}&no_urut={no_urut}"
                
                try:
                    self.logger.info(f"[{i}] Getting bjdt_id for Doc: {doc_no}, No Urut: {no_urut}...")
                    response = requests.get(url)
                    response.raise_for_status()
                    data = response.json()
                    bjdt_id = data.get("bjdt_id", "")
                    
                    if not bjdt_id:
                        self.logger.warning(f"bjdt_id NOT FOUND")
                        error_logs.append(f"bjdt_id not found for {doc_no} no_urut {no_urut}")
                    else:
                        self.logger.info(f"Found bjdt_id: {bjdt_id}")
                    
                    row_write[index_bjdt_id].value = bjdt_id
                    
                except Exception as e:
                    self.logger.error(f"ERROR on {url}: {e}")
                    error_logs.append(f"ERROR on {url}: {e}")
                    # row_write cells can be MergedCell instances; runtime assignment is OK
                    row_write[index_bjdt_id].value = ""  # type: ignore
        
        # Save workbook
        wb_write.save(self.file_excel_2.get())
        
        # Save error logs
        if error_logs:
            with open("bjdt_error_log.txt", mode="w", encoding="utf-8") as error_file:
                for log in error_logs:
                    error_file.write(log + "\n")
            self.logger.warning(f"Saved {len(error_logs)} errors to bjdt_error_log.txt")
        
        self.logger.info("BJDT ID process completed")
        
    def _generate_csv_files_logic(self):
        """Generate CSV files separated by date"""
        # Calculate formulas
        with xw.App(visible=False) as app:
            wb = app.books.open(self.file_excel_2.get())
            wb.app.calculate()
            wb.save()
            wb.close()
        
        # Read Excel file
        df = pd.read_excel(self.file_excel_2.get(), sheet_name=self.nama_sheet_3.get(), engine='openpyxl')
        
        # Convert date column to datetime
        df['tglrencanauji'] = pd.to_datetime(df['tglrencanauji'], errors='coerce')
        
        # Create output directory
        os.makedirs(self.output_dir.get(), exist_ok=True)
        
        # Group by date and save as CSV
        for tgl, group in df.groupby(df['tglrencanauji'].dt.date):
            output_filename = os.path.join(self.output_dir.get(), f"{tgl}.csv")
            group.to_csv(output_filename, index=False, sep=';', encoding='utf-8-sig')
            self.logger.info(f"Saved: {output_filename}")
        # Clear ODOO sheet
        wb = openpyxl.load_workbook(self.file_excel_2.get())
        ws = wb[self.nama_sheet_2.get()]
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.value = None
        
        wb.save(self.file_excel_2.get())
        # Clear bjdt_id sheet (columns 3 and beyond)
        ws = wb[self.nama_sheet_4.get()]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
            for cell in row:
                cell.value = None
        
        wb.save(self.file_excel_2.get())
        wb.close()
        
        self.logger.info("CSV files generated and sheets cleared")
        
    def upload_to_database(self):
        """Upload CSV data to PostgreSQL database"""
        today = datetime.now().strftime('%Y-%m-%d')
        csv_file = os.path.join(self.output_dir.get(), f'{today}.csv')
        
        if not os.path.isfile(csv_file):
            self.logger.error(f"File {csv_file} does not exist.")
            return
        
        # Create database connection
        engine = create_engine(f'postgresql+psycopg2://{self.db_user.get()}:{self.db_password.get()}@{self.db_host.get()}:{self.db_port.get()}/{self.db_name.get()}')
        
        # Read CSV
        df = pd.read_csv(csv_file, delimiter=';')
        self.logger.info(f"Number of rows to upload: {len(df)}")
        
        # Upload to PostgreSQL
        try:
            df.to_sql(self.table_name.get(), engine, if_exists='append', index=False)
            self.logger.info("Upload completed successfully")
            
            # Display uploaded data
            try:
                uploaded_df = pd.read_sql_table(self.table_name.get(), engine)
                self.logger.info(f"Total records in database: {len(uploaded_df)}")
            except Exception as e:
                self.logger.error(f"Failed to display uploaded data: {e}")
                
        except Exception as e:
            # Some DB exceptions (like SQLAlchemy DBAPIError) expose an 'orig' with pgcode.
            # Access it safely to avoid static type errors on generic Exception.
            orig = getattr(e, 'orig', None)
            pgcode = getattr(orig, 'pgcode', None)
            if pgcode == '23505':
                self.logger.error("Upload failed: DUPLICATE DATA DETECTED")
            else:
                self.logger.error(f"Upload error: {e}")

    def upload_to_database_files(self):
        """Upload CSV data to PostgreSQL database"""
        today = datetime.now().strftime('%Y-%m-%d')
        csv_file = os.path.join(self.file_csv.get())
        
        if not os.path.isfile(csv_file):
            self.logger.error(f"File {csv_file} does not exist.")
            return
        
        # Create database connection
        engine = create_engine(f'postgresql+psycopg2://{self.db_user.get()}:{self.db_password.get()}@{self.db_host.get()}:{self.db_port.get()}/{self.db_name.get()}')
        
        # Read CSV
        df = pd.read_csv(csv_file, delimiter=';')
        self.logger.info(f"Number of rows to upload: {len(df)}")
        
        # Upload to PostgreSQL
        try:
            df.to_sql(self.table_name.get(), engine, if_exists='append', index=False)
            self.logger.info("Upload completed successfully")
            
            # Display uploaded data
            try:
                uploaded_df = pd.read_sql_table(self.table_name.get(), engine)
                self.logger.info(f"Total records in database: {len(uploaded_df)}")
            except Exception as e:
                self.logger.error(f"Failed to display uploaded data: {e}")
                
        except Exception as e:
            orig = getattr(e, 'orig', None)
            pgcode = getattr(orig, 'pgcode', None)
            if pgcode == '23505':
                self.logger.error("Upload failed: DUPLICATE DATA DETECTED")
            else:
                self.logger.error(f"Upload error: {e}")

    def toggle_sync(self):
        """Toggle ODOO sync daemon on/off"""
        if not self.daemon_running:
            self.continue_to_sync()
        else:
            self.stop_sync()

    def continue_to_sync(self):
        """Continue to ODOO sync with thread-safe implementation"""
        # Prevent multiple instances
        if self.daemon_running:
            messagebox.showwarning("Warning", "Sync is already running")
            return
        
        # Run initialization in separate thread to avoid blocking
        threading.Thread(target=self._init_sync, daemon=True).start()
    
    def _init_sync(self):
        """Initialize sync in background thread"""
        try:
            # Import daemon (safe in worker thread)
            from modules.daemon_sync import threadSinkData
            
            # Read config (I/O operation in worker thread)
            fileConfig = resource_path("config.cnf")
            config = configparser.RawConfigParser()
            config.read(fileConfig)
            tunda = config.get("daemon", "delay") if config.has_option("daemon", "delay") else '2'
            
            # Update GUI from main thread
            self.root.after(0, self._update_sync_ui_starting)
            
            # Create daemon with thread-safe log_queue (NOT log_widget)
            self.daemon = threadSinkData(
                1, 
                "ThreadSinkron", 
                float(tunda),
                log_queue=self.log_queue  # Pass queue, NOT widget
            )
            self.daemon.start()
            self.daemon_running = True
            
            # Log success from main thread
            self.root.after(0, self._log_sync_started)
            
        except Exception as e:
            error_msg = f"Error starting ODOO sync: {str(e)}"
            # Show error in main thread
            self.root.after(0, lambda: self._show_sync_error(error_msg))
    
    def _update_sync_ui_starting(self):
        """Update UI when sync starts (main thread only)"""
        self.progress.start()
        self.sync_btn.configure(
            state='normal',
            text='⏹️ Stop ODOO Sync',
            fg_color="#DC3545",
            hover_color="#C82333"
        )
    
    def _log_sync_started(self):
        """Log sync started messages (main thread)"""
        if self.logger:
            self.logger.info("=" * 60)
            self.logger.info("ODOO sync daemon started successfully")
            self.logger.info(f"Thread ID: {self.daemon.ident if self.daemon else 'N/A'}")
            self.logger.info(f"Thread name: {self.daemon.name if self.daemon else 'N/A'}")
            self.logger.info(f"Is alive: {self.daemon.is_alive() if self.daemon else False}")
            self.logger.info("=" * 60)
        
    
    def _show_sync_error(self, error_msg):
        """Show error message (main thread only)"""
        self.daemon_running = False
        if self.logger:
            self.logger.error(error_msg)
        messagebox.showerror("Error", error_msg)
        
        # Reset UI
        self.sync_btn.configure(
            state='normal',
            text='🔄 Start ODOO Sync',
            fg_color="#28A745",
            hover_color="#218838"
        )

    def stop_sync(self):
        """Stop the ODOO sync daemon (thread-safe)"""
        if not self.daemon_running:
            return
        
        def _stop_daemon():
            try:
                if self.daemon and self.daemon.is_alive():
                    self.daemon.stop()
                    self.daemon.join(timeout=5.0)
                    
                    if self.daemon.is_alive():
                        msg = "Daemon did not stop within timeout"
                        self.root.after(0, lambda: self.logger.warning(msg))
                    else:
                        msg = "Daemon stopped successfully"
                        self.root.after(0, lambda: self.logger.info(msg))
                
                # Update UI in main thread
                self.root.after(0, self._update_sync_ui_stopped)
                
            except Exception as e:
                error_msg = f"Error stopping sync: {str(e)}"
                self.root.after(0, lambda: self._show_sync_error(error_msg))
        
        # Stop in background thread
        threading.Thread(target=_stop_daemon, daemon=True).start()
    
    def _update_sync_ui_stopped(self):
        """Update UI when sync stops (main thread only)"""
        self.daemon_running = False
        self.daemon = None
        self.progress.stop()
        self.sync_btn.configure(
            state='normal',
            text='🔄 Start ODOO Sync',
            fg_color="#28A745",
            hover_color="#218838"
        )
        if self.logger:
            self.logger.info("ODOO sync daemon stopped")

    def run_daemon(self):
        # Placeholder for daemon logic - REMOVED as we use _init_sync now
        pass

    def test_db_connection(self):
        try:
            engine = create_engine(f'postgresql+psycopg2://{self.db_user.get()}:{self.db_password.get()}@{self.db_host.get()}:{self.db_port.get()}/{self.db_name.get()}')
            connection = engine.connect()
            connection.close()
            messagebox.showinfo("Success", "Database connection successful!")
        except Exception as e:
            messagebox.showerror("Error", f"Database connection failed: {str(e)}")

