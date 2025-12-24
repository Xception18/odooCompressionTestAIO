import pandas as pd
import logging
import openpyxl
import random
from pathlib import Path
from typing import Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)

class ExcelDataProcessor:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.data = None
        self.load_excel_data()
    
    def load_excel_data(self):
        """Load data from Excel file"""
        try:
            self.data = pd.read_excel(self.excel_file_path)
            logger.info(f"Loaded Excel file with {len(self.data)} rows")
            logger.info(f"Columns: {list(self.data.columns)}")
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise
    
    def get_row_data(self, row_index):
        """Get data for specific row"""
        if self.data is None or row_index >= len(self.data):
            return None
        return self.data.iloc[row_index]
    
    def should_duplicate(self, current_row_index):
        """Check if next row has same kode_benda_uji and proyek"""
        if self.data is None or current_row_index >= len(self.data) - 1:
            return False
        
        current_row = self.get_row_data(current_row_index)
        next_row = self.get_row_data(current_row_index + 1)
        
        if current_row is None or next_row is None:
            return False
        
        # Get kode_benda_uji from column 3 (index 2)
        current_kode = str(current_row.iloc[2]) if len(current_row) > 2 else ""
        next_kode = str(next_row.iloc[2]) if len(next_row) > 2 else ""
        
        # Get proyek from column 4 (index 3)
        current_proyek = str(current_row.iloc[3]) if len(current_row) > 3 else ""
        next_proyek = str(next_row.iloc[3]) if len(next_row) > 3 else ""
        
        # Check if both values are the same
        same_kode = current_kode == next_kode
        same_proyek = current_proyek == next_proyek
        
        logger.info(f"Duplicate check - Current row {current_row_index + 1}:")
        logger.info(f"  Current kode_benda_uji: {current_kode}")
        logger.info(f"  Next kode_benda_uji: {next_kode}")
        logger.info(f"  Current proyek: {current_proyek}")
        logger.info(f"  Next proyek: {next_proyek}")
        logger.info(f"  Same kode: {same_kode}, Same proyek: {same_proyek}")
        
        return same_kode and same_proyek


class ExcelBebanProcessor:
    """
    Processor untuk menghitung nilai BEBAN pada file Excel
    berdasarkan MUTU, KODE_BENDA_UJI, UMUR, dan DOCKET
    """
    
    def __init__(self, file_path: str, sheet_name: str = 'ODOO'):
        """
        Inisialisasi processor
        
        Args:
            file_path: Path ke file Excel
            sheet_name: Nama sheet yang akan diproses (default: 'ODOO')
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None
        
        # Tracking untuk logika umur 28
        self.previous_docket: Optional[str] = None
        self.previous_beban: Optional[float] = None
        self.previous_mutu: Optional[str] = None
        
        # Mapping kolom (1-based index untuk Excel)
        self.column_map = {
            'DOCKET': 1,      # Kolom A -> index 1
            'MUTU': 6,        # Kolom F -> index 6
            'UMUR': 9,        # Kolom I -> index 9
            'KODE_BENDA_UJI': 16,  # Kolom P -> index 16
            'BEBAN': 18       # Kolom R -> index 18
        }
        
        # Keywords untuk pengecekan
        self.keywords = ["PP - TOL PTB", "WASKITA - ABP JO", "HK - JAKON JO", "WIKA - ADHI JO"]
    
    def load_excel(self) -> bool:
        """Membuka file Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook[self.sheet_name]
            print(f"✓ File Excel berhasil dibuka: {self.file_path}")
            print(f"✓ Sheet aktif: {self.sheet_name}")
            return True
        except FileNotFoundError:
            print(f"✗ Error: File tidak ditemukan - {self.file_path}")
            return False
        except KeyError:
            print(f"✗ Error: Sheet '{self.sheet_name}' tidak ditemukan")
            return False
        except Exception as e:
            print(f"✗ Error saat membuka file: {str(e)}")
            return False
    
    def get_cell_value(self, row_index: int, column_key: str):
        """Mendapatkan nilai cell berdasarkan row dan column key"""
        if self.sheet is None:
            raise RuntimeError("Sheet belum dimuat. Jalankan load_excel() terlebih dahulu.")
        
        col_index = self.column_map[column_key]
        return self.sheet.cell(row=row_index, column=col_index).value
    
    def set_cell_value(self, row_index: int, column_key: str, value: str) -> None:
        """Mengisi nilai cell berdasarkan row dan column key"""
        if self.sheet is None:
            raise RuntimeError("Sheet belum dimuat. Jalankan load_excel() terlebih dahulu.")
        
        col_index = self.column_map[column_key]
        self.sheet.cell(row=row_index, column=col_index, value=value)
    
    def calculate_beban(self, row_index: int) -> Optional[str]:
        """
        Menghitung nilai BEBAN untuk baris tertentu
        
        Args:
            row_index: Index baris di Excel (1-based)
            
        Returns:
            String nilai BEBAN atau None jika tidak memenuhi kriteria
        """
        # Ambil data dari row
        mutu = str(self.get_cell_value(row_index, 'MUTU') or '')
        kode_benda_uji = str(self.get_cell_value(row_index, 'KODE_BENDA_UJI') or '')
        umur = self.get_cell_value(row_index, 'UMUR')
        docket = self.get_cell_value(row_index, 'DOCKET')


        if "Class B-2" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(390.12, 460.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(580, 700)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = 645
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-19, -8)  # Turun
                    else:
                        adjustment = random.uniform(8, 19)  # Naik
                    new_beban = max(520, min(self.previous_beban + adjustment, 750))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = 645
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-84, -68)  # Turun signifikan
                    else:
                        adjustment = random.uniform(74, 96)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(590, min(new_beban, 750))

                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(580, 700)
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(580, 700)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "K-400" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(390.12, 460.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(580, 700)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = 645
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-19, -8)  # Turun
                    else:
                        adjustment = random.uniform(8, 19)  # Naik
                    new_beban = max(520, min(self.previous_beban + adjustment, 750))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = 645
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-84, -68)  # Turun signifikan
                    else:
                        adjustment = random.uniform(74, 96)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(590, min(new_beban, 750))

                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(580, 700)
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(580, 700)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"

        elif "Class B-1" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(360.12, 420.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(540, 660)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = 600
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-18, -6)  # Turun
                    else:
                        adjustment = random.uniform(5, 15)  # Naik
                    new_beban = max(500, min(self.previous_beban + adjustment, 700))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket:
                    mean = 625
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-83, -71)  # Turun signifikan
                    else:
                        adjustment = random.uniform(59, 71)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(540, min(new_beban, 710))
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(540, 660)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"

        elif "Fc-30" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(360.12, 420.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(540, 660)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = 600
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-18, -6)  # Turun
                    else:
                        adjustment = random.uniform(5, 15)  # Naik
                    new_beban = max(500, min(self.previous_beban + adjustment, 700))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket:
                    mean = 625
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-83, -71)  # Turun signifikan
                    else:
                        adjustment = random.uniform(59, 71)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(540, min(new_beban, 710))
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(540, 660)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "K-350" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(328.12, 380.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(510, 640)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (510 + 640) / 2  # 575
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-21, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 19)  # Naik
                    new_beban = max(510, min(self.previous_beban + adjustment, 640))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = (510 + 640) / 2  # 575
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-66, -52)  # Turun signifikan
                    else:
                        adjustment = random.uniform(62, 74)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(510, min(new_beban, 640))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(510, 640)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(510, 640)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"

        elif "K-300" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(290.12, 360.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(430, 500)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (430 + 500) / 2  # 465
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-13, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 13)  # Naik
                    new_beban = max(430, min(self.previous_beban + adjustment, 500))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = 475
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-74, -67)  # Turun signifikan
                    else:
                        adjustment = random.uniform(59, 76)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(398, min(new_beban, 560))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(430, 500)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(430, 500)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Fc-25" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(290.12, 360.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(430, 500)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (430 + 500) / 2  # 465
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(430, min(self.previous_beban + adjustment, 500))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = (430 + 500) / 2  # 465
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-34, -20)  # Turun signifikan
                    else:
                        adjustment = random.uniform(20, 34)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(430, min(new_beban, 500))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(430, 500)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(430, 500)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "K-250" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(230.12, 285.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(365, 475)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = 420
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-10, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 10)  # Naik
                    new_beban = max(365, min(self.previous_beban + adjustment, 475))
                # Skenario 3: Docket berbeda dan mutu sama
                elif self.previous_docket != docket:
                    mean = 400
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-89, -76)  # Turun signifikan
                    else:
                        adjustment = random.uniform(66, 75)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(300, min(new_beban, 474))                  
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(365, 475)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Fc-20" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(230.12, 285.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(365, 475)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = 420
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-10, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 10)  # Naik
                    new_beban = max(365, min(self.previous_beban + adjustment, 475))
                # Skenario 3: Docket berbeda dan mutu sama
                elif self.previous_docket != docket:
                    mean = 400
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-89, -76)  # Turun signifikan
                    else:
                        adjustment = random.uniform(66, 75)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(300, min(new_beban, 474))                  
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(365, 475)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Class C" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(230.12, 285.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(365, 475)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = 420
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-10, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 10)  # Naik
                    new_beban = max(365, min(self.previous_beban + adjustment, 475))
                # Skenario 3: Docket berbeda dan mutu sama
                elif self.previous_docket != docket:
                    mean = 400
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-89, -76)  # Turun signifikan
                    else:
                        adjustment = random.uniform(66, 75)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(300, min(new_beban, 474))                  
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(365, 475)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "K-175" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(165.12, 215.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(255, 305)
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(255, min(self.previous_beban + adjustment, 305))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-24, -8)  # Turun signifikan
                    else:
                        adjustment = random.uniform(8, 24)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(255, min(new_beban, 305))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(255, 305)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(255, 305)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Fc-15" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(165.12, 215.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(255, 305)
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(255, min(self.previous_beban + adjustment, 305))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-24, -8)  # Turun signifikan
                    else:
                        adjustment = random.uniform(8, 24)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(255, min(new_beban, 305))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(255, 305)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(255, 305)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        elif "Class D" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(165.12, 215.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(255, 305)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-29, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 29)  # Naik
                    new_beban = max(255, min(self.previous_beban + adjustment, 305))
                # Skenario 3: Docket berbeda & mutu sama
                elif self.previous_docket != docket and self.previous_mutu == mutu:
                    mean = (255 + 305) / 2  # 280
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-24, -8)  # Turun signifikan
                    else:
                        adjustment = random.uniform(8, 24)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(255, min(new_beban, 305))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(255, 305)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(255, 305)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"

        elif "K-125" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(120.12, 165.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(180, 280)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (180 + 280) / 2
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-11, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 11)  # Naik
                    new_beban = max(180, min(self.previous_beban + adjustment, 280))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (180 + 280) / 2
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-48, -38)  # Turun signifikan
                    else:
                        adjustment = random.uniform(38, 48)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(180, min(new_beban, 280))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(180, 280)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(180, 280)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
            
        elif "Fc-10" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(120.12, 165.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(180, 280)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (180 + 280) / 2
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-11, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 11)  # Naik
                    new_beban = max(180, min(self.previous_beban + adjustment, 280))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (180 + 280) / 2
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-48, -38)  # Turun signifikan
                    else:
                        adjustment = random.uniform(38, 48)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(180, min(new_beban, 280))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(180, 280)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(180, 280)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"

        elif "Class E-1" in mutu:
        # Logika perhitungan berdasarkan UMUR
            if umur == 7:
                random_beban = random.uniform(120.12, 165.34)
                return f"{random_beban:.2f}"        
            elif umur == 28:
                # Umur 28 hari dengan logika khusus
                # Skenario 1: Kondisi awal (docket & beban sebelumnya = 0.0 atau None)
                if (self.previous_docket is None or self.previous_docket == 0.0 or 
                    self.previous_beban is None or self.previous_beban == 0.0):
                    new_beban = random.uniform(180, 280)
                
                # Skenario 2: Docket sama & beban != 0
                elif self.previous_docket == docket and self.previous_beban != None:
                    mean = (180 + 280) / 2
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-11, -3)  # Turun
                    else:
                        adjustment = random.uniform(3, 11)  # Naik
                    new_beban = max(180, min(self.previous_beban + adjustment, 280))
                # Skenario 3: Docket berbeda
                elif self.previous_docket != docket:
                    mean = (180 + 280) / 2
                    if self.previous_beban >= mean:
                        adjustment = random.uniform(-48, -38)  # Turun signifikan
                    else:
                        adjustment = random.uniform(38, 48)  # Naik signifikan
                    new_beban = self.previous_beban + adjustment
                    # Pastikan tetap dalam range
                    new_beban = max(180, min(new_beban, 280))
                
                # Skenario 4: Docket berbeda & mutu berbeda
                elif self.previous_docket != docket and self.previous_mutu != mutu:
                    new_beban = random.uniform(180, 280)
                    
                else:
                    # Fallback (seharusnya tidak terjadi)
                    new_beban = random.uniform(180, 280)
                # Update tracking variables
                self.previous_docket = str(docket) if docket is not None else None
                self.previous_beban = new_beban
                
                return f"{new_beban:.2f}"
            else:
                # Umur selain 7 atau 28
                return "0.00"
        
        else:
            # Default fallback jika mutu tidak dikenali
            return "0.00"
    



    def process_all_rows(self, start_row: int = 2) -> None:
        """
        Memproses semua baris di Excel
        
        Args:
            start_row: Baris awal data (default: 2, asumsi baris 1 adalah header)
        """
        if self.sheet is None:
            raise RuntimeError("Sheet belum dimuat. Jalankan load_excel() terlebih dahulu.")
        
        max_row = self.sheet.max_row
        print(f"Memulai proses perhitungan untuk {max_row - start_row + 1} baris...")
        
        count = 0
        for row in range(start_row, max_row + 1):
            beban = self.calculate_beban(row)
            if beban:
                self.set_cell_value(row, 'BEBAN', beban)
                count += 1
                
        print(f"✓ Selesai! {count} baris telah diperbarui.")
    
    def save_excel(self, output_path: Optional[str] = None) -> bool:
        """
        Menyimpan file Excel
        
        Args:
            output_path: Path tujuan simpan (default: overwrite file asli)
        """
        if self.workbook is None:
            return False
            
        target_path = output_path or self.file_path
        try:
            self.workbook.save(target_path)
            print(f"✓ File berhasil disimpan ke: {target_path}")
            return True
        except Exception as e:
            print(f"✗ Error saat menyimpan file: {str(e)}")
            return False
            
    def close(self):
        """Menutup workbook"""
        if self.workbook:
            self.workbook.close()
